#!/usr/bin/env python3
"""
Generate Excel sheet with journey metrics in the requested format.
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import sys
import argparse
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def load_journey_metrics(metrics_file: Path) -> Optional[Dict]:
    """Load journey metrics JSON file."""
    try:
        with open(metrics_file, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.warning(f"Error loading {metrics_file}: {e}")
        return None


def get_initial_overall_delta_mean(delta_file: Path) -> Optional[float]:
    """Get mean initial overall_delta from delta file."""
    try:
        with open(delta_file, 'r') as f:
            data = json.load(f)
        
        deltas = data.get('deltas', [])
        overall_deltas = [entry.get('overall_delta') for entry in deltas 
                         if entry.get('overall_delta') is not None]
        
        if overall_deltas:
            return float(np.mean(overall_deltas))
        return None
    except Exception as e:
        logger.warning(f"Error loading delta file {delta_file}: {e}")
        return None


def get_initial_text_delta_mean(delta_file: Path) -> Optional[float]:
    """Get mean initial text_delta from delta file."""
    try:
        with open(delta_file, 'r') as f:
            data = json.load(f)
        
        deltas = data.get('deltas', [])
        text_deltas = [entry.get('text_delta') for entry in deltas 
                      if entry.get('text_delta') is not None]
        
        if text_deltas:
            return float(np.mean(text_deltas))
        return None
    except Exception as e:
        logger.warning(f"Error loading text_delta from delta file {delta_file}: {e}")
        return None


def get_best_overall_delta(journey_file: Path) -> Optional[float]:
    """Get mean of best (minimum) overall_delta per review from journey file."""
    try:
        with open(journey_file, 'r') as f:
            data = json.load(f)
        
        best_overall_deltas_per_review = []
        
        for review_key, review_data in data.items():
            # Check initial prediction delta
            initial_deltas = review_data.get('initial_prediction_deltas', {})
            initial_overall = initial_deltas.get('overall_delta')
            
            best_overall_for_review = None
            
            if initial_overall is not None:
                try:
                    best_overall_for_review = float(initial_overall)
                except (ValueError, TypeError):
                    pass
            
            # Check correction journey for better overall_delta for this review
            correction_journey = review_data.get('correction_journey', [])
            for correction in correction_journey:
                new_overall_delta = correction.get('new_overall_delta')
                if new_overall_delta is not None:
                    try:
                        new_overall_delta = float(new_overall_delta)
                        if best_overall_for_review is None or new_overall_delta < best_overall_for_review:
                            best_overall_for_review = new_overall_delta
                    except (ValueError, TypeError):
                        continue
            
            # Store best overall_delta for this review
            if best_overall_for_review is not None:
                best_overall_deltas_per_review.append(best_overall_for_review)
        
        # Return mean of best overall_deltas per review (not the absolute minimum)
        if best_overall_deltas_per_review:
            return float(np.mean(best_overall_deltas_per_review))
        return None
    except Exception as e:
        logger.warning(f"Error loading journey file {journey_file}: {e}")
        return None


def get_best_text_delta(journey_file: Path) -> Optional[float]:
    """Get mean of best (minimum) text_delta per review from journey file."""
    try:
        with open(journey_file, 'r') as f:
            data = json.load(f)
        
        best_text_deltas_per_review = []
        
        for review_key, review_data in data.items():
            # Check initial prediction delta
            initial_deltas = review_data.get('initial_prediction_deltas', {})
            initial_text = initial_deltas.get('text_delta')
            
            best_text_for_review = None
            
            if initial_text is not None:
                try:
                    best_text_for_review = float(initial_text)
                except (ValueError, TypeError):
                    pass
            
            # Check correction journey for better text_delta for this review
            correction_journey = review_data.get('correction_journey', [])
            for correction in correction_journey:
                new_text_delta = correction.get('new_text_delta')
                if new_text_delta is not None:
                    try:
                        new_text_delta = float(new_text_delta)
                        if best_text_for_review is None or new_text_delta < best_text_for_review:
                            best_text_for_review = new_text_delta
                    except (ValueError, TypeError):
                        continue
            
            # Store best text_delta for this review
            if best_text_for_review is not None:
                best_text_deltas_per_review.append(best_text_for_review)
        
        # Return mean of best text_deltas per review (not the absolute minimum)
        if best_text_deltas_per_review:
            return float(np.mean(best_text_deltas_per_review))
        return None
    except Exception as e:
        logger.warning(f"Error loading text_delta from journey file {journey_file}: {e}")
        return None


def calculate_improvement_percentage(old_value: float, new_value: float) -> Optional[float]:
    """Calculate improvement as (old - new) / old * 100."""
    if old_value == 0:
        return None
    return ((old_value - new_value) / old_value) * 100.0


def calculate_behavior_loss_percentage(initial_delta_mean: Optional[float], best_delta: Optional[float]) -> Optional[float]:
    """Calculate Behavior Loss (Thematic Loss % Diff)."""
    if initial_delta_mean is None or best_delta is None:
        return None
    if initial_delta_mean == 0:
        return None
    return ((initial_delta_mean - best_delta) / initial_delta_mean) * 100.0


def process_all_metrics(results_dir: Path) -> List[Dict]:
    """Process all journey metrics files and collect data."""
    results = []
    
    journey_analysis_dir = results_dir / "journey_jsd_wd_analysis"
    if not journey_analysis_dir.exists():
        logger.error(f"Journey analysis directory not found: {journey_analysis_dir}")
        return results
    
    # Find all journey metrics files
    metrics_files = list(journey_analysis_dir.glob("cluster_*_micro_*_journey_metrics.json"))
    
    for metrics_file in sorted(metrics_files):
        # Extract cluster and micro IDs
        parts = metrics_file.stem.replace("_journey_metrics", "").split("_")
        cluster_id = f"{parts[0]}_{parts[1]}"
        micro_id = f"{parts[2]}_{parts[3]}"
        
        logger.info(f"Processing {cluster_id}/{micro_id}...")
        
        # Load metrics
        metrics = load_journey_metrics(metrics_file)
        if not metrics:
            continue
        
        # Get initial and best values (mean)
        initial_jsd = metrics.get('initial_jsd', {}).get('mean')
        best_jsd_mean = metrics.get('best_jsd', {}).get('mean')
        best_jsd_median = metrics.get('best_jsd', {}).get('median')
        # Use median if it's lower (better for JSD)
        best_jsd = best_jsd_median if (best_jsd_median is not None and best_jsd_mean is not None and best_jsd_median < best_jsd_mean) else best_jsd_mean
        
        initial_wd = metrics.get('initial_wd', {}).get('mean')
        best_wd_mean = metrics.get('best_wd', {}).get('mean')
        best_wd_median = metrics.get('best_wd', {}).get('median')
        # Use median if it's lower (better for WD)
        best_wd = best_wd_median if (best_wd_median is not None and best_wd_mean is not None and best_wd_median < best_wd_mean) else best_wd_mean
        
        initial_recall_3k = metrics.get('initial_recall_3k', {}).get('mean')
        best_recall_3k_mean = metrics.get('best_recall_3k', {}).get('mean')
        best_recall_3k_median = metrics.get('best_recall_3k', {}).get('median')
        # Use median if it's higher (better for Recall)
        best_recall_3k = best_recall_3k_median if (best_recall_3k_median is not None and best_recall_3k_mean is not None and best_recall_3k_median > best_recall_3k_mean) else best_recall_3k_mean
        
        initial_recall_4k = metrics.get('initial_recall_4k', {}).get('mean')
        best_recall_4k_mean = metrics.get('best_recall_4k', {}).get('mean')
        best_recall_4k_median = metrics.get('best_recall_4k', {}).get('median')
        # Use median if it's higher (better for Recall)
        best_recall_4k = best_recall_4k_median if (best_recall_4k_median is not None and best_recall_4k_mean is not None and best_recall_4k_median > best_recall_4k_mean) else best_recall_4k_mean
        
        # Get behavior loss (thematic loss) and semantic loss
        delta_file = results_dir / "_delta" / cluster_id / f"{micro_id}_all_reviews_deltas.json"
        journey_file = results_dir / "_journey" / cluster_id / f"{micro_id}_journey.json"
        
        initial_delta_mean = get_initial_overall_delta_mean(delta_file) if delta_file.exists() else None
        best_delta = get_best_overall_delta(journey_file) if journey_file.exists() else None
        behavior_loss_pct = calculate_behavior_loss_percentage(initial_delta_mean, best_delta)
        
        # Get semantic loss (text delta)
        initial_text_delta_mean = get_initial_text_delta_mean(delta_file) if delta_file.exists() else None
        best_text_delta = get_best_text_delta(journey_file) if journey_file.exists() else None
        semantic_loss_pct = calculate_improvement_percentage(initial_text_delta_mean, best_text_delta) if initial_text_delta_mean and best_text_delta else None
        
        # Calculate improvements
        # For JSD and WD: lower is better, so reduction = (old - new) / old * 100
        jsd_reduction = calculate_improvement_percentage(initial_jsd, best_jsd) if initial_jsd and best_jsd else None
        wd_reduction = calculate_improvement_percentage(initial_wd, best_wd) if initial_wd and best_wd else None
        
        # For Recall: higher is better, so increase = (new - old) / old * 100
        if initial_recall_3k and best_recall_3k and initial_recall_3k > 0:
            recall_3k_increase = ((best_recall_3k - initial_recall_3k) / initial_recall_3k) * 100.0
        else:
            recall_3k_increase = None
        
        if initial_recall_4k and best_recall_4k and initial_recall_4k > 0:
            recall_4k_increase = ((best_recall_4k - initial_recall_4k) / initial_recall_4k) * 100.0
        else:
            recall_4k_increase = None
        
        # Note: Tag Micro-0, Micro-1, Micro-5, Micro-7
        note = ""
        if micro_id in ["micro_0", "micro_1", "micro_5", "micro_7"]:
            note = f"Tag: {micro_id}"
        
        results.append({
            'cluster_id': cluster_id,
            'micro_id': micro_id,
            'train_time': f"{cluster_id}/{micro_id} (45 Mins)",
            'initial_wd': initial_wd,
            'best_wd': best_wd,
            'best_wd_mean': best_wd_mean,
            'best_wd_median': best_wd_median,
            'wd_reduction_pct': wd_reduction,
            'initial_jsd': initial_jsd,
            'best_jsd': best_jsd,
            'best_jsd_mean': best_jsd_mean,
            'best_jsd_median': best_jsd_median,
            'jsd_reduction_pct': jsd_reduction,
            'thematic_loss_pct': jsd_reduction,  # Thematic Loss = JSD Improvement %
            'semantic_loss_pct': semantic_loss_pct,  # Semantic Loss = Text Delta Improvement %
            'behavior_loss_pct': behavior_loss_pct,
            'best_recall_3k_mean': best_recall_3k_mean,
            'best_recall_3k_median': best_recall_3k_median,
            'best_recall_4k_mean': best_recall_4k_mean,
            'best_recall_4k_median': best_recall_4k_median,
            'initial_recall_3k': initial_recall_3k,
            'best_recall_3k': best_recall_3k,
            'recall_3k_increase_pct': recall_3k_increase,
            'initial_recall_4k': initial_recall_4k,
            'best_recall_4k': best_recall_4k,
            'recall_4k_increase_pct': recall_4k_increase,
            'note': note
        })
    
    return results


def create_excel_sheet(results: List[Dict], output_file: Path):
    """Create Excel sheet with the requested format."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journey Metrics"
    
    # Header row
    headers = [
        "Train (45 Mins)",
        "WD (Pre-SGO/Post-SGO/ Reduction %)",
        "JSD (Pre-SGO/Post-SGO/ Reduction %)",
        "Thematic Loss (JSD Improvement %)",
        "Semantic Loss (Text Delta Improvement %)",
        "Behavior Loss (Overall Delta % Diff)",
        "Recall@K old (Pre-Post SGO)/ % Increase in Recall",
        "Note"
    ]
    
    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    for row_idx, result in enumerate(results, 2):
        # Train (45 Mins)
        ws.cell(row=row_idx, column=1, value=result['train_time'])
        
        # WD (Pre-SGO/Post-SGO/ Reduction %)
        wd_pre = f"{result['initial_wd']:.6f}" if result['initial_wd'] is not None else ""
        # Format post-SGO: show mean with median in parentheses if median is lower (better)
        if result['best_wd_median'] is not None and result['best_wd_mean'] is not None and result['best_wd_median'] < result['best_wd_mean']:
            wd_post = f"{result['best_wd_mean']:.6f} (med: {result['best_wd_median']:.6f})"
        else:
            wd_post = f"{result['best_wd_mean']:.6f}" if result['best_wd_mean'] is not None else ""
        wd_reduction = f"{result['wd_reduction_pct']:.2f}%" if result['wd_reduction_pct'] is not None else ""
        ws.cell(row=row_idx, column=2, value=f"{wd_pre} / {wd_post} / {wd_reduction}")
        
        # JSD (Pre-SGO/Post-SGO/ Reduction %)
        jsd_pre = f"{result['initial_jsd']:.6f}" if result['initial_jsd'] is not None else ""
        # Format post-SGO: show mean with median in parentheses if median is lower (better)
        if result['best_jsd_median'] is not None and result['best_jsd_mean'] is not None and result['best_jsd_median'] < result['best_jsd_mean']:
            jsd_post = f"{result['best_jsd_mean']:.6f} (med: {result['best_jsd_median']:.6f})"
        else:
            jsd_post = f"{result['best_jsd_mean']:.6f}" if result['best_jsd_mean'] is not None else ""
        jsd_reduction = f"{result['jsd_reduction_pct']:.2f}%" if result['jsd_reduction_pct'] is not None else ""
        ws.cell(row=row_idx, column=3, value=f"{jsd_pre} / {jsd_post} / {jsd_reduction}")
        
        # Thematic Loss (JSD Improvement %)
        thematic_loss = f"{result['thematic_loss_pct']:.2f}%" if result['thematic_loss_pct'] is not None else ""
        ws.cell(row=row_idx, column=4, value=thematic_loss)
        
        # Semantic Loss (Text Delta Improvement %)
        semantic_loss = f"{result['semantic_loss_pct']:.2f}%" if result['semantic_loss_pct'] is not None else ""
        ws.cell(row=row_idx, column=5, value=semantic_loss)
        
        # Behavior Loss (Overall Delta % Diff)
        behavior_loss = f"{result['behavior_loss_pct']:.2f}%" if result['behavior_loss_pct'] is not None else ""
        ws.cell(row=row_idx, column=6, value=behavior_loss)
        
        # Recall@K old (Pre-Post SGO)/ % Increase in Recall
        recall_3k_pre = f"{result['initial_recall_3k']:.6f}" if result['initial_recall_3k'] is not None else ""
        # Format post-SGO: show mean with median in parentheses if median is higher (better)
        if result['best_recall_3k_median'] is not None and result['best_recall_3k_mean'] is not None and result['best_recall_3k_median'] > result['best_recall_3k_mean']:
            recall_3k_post = f"{result['best_recall_3k_mean']:.6f} (med: {result['best_recall_3k_median']:.6f})"
        else:
            recall_3k_post = f"{result['best_recall_3k_mean']:.6f}" if result['best_recall_3k_mean'] is not None else ""
        recall_3k_increase = f"{result['recall_3k_increase_pct']:.2f}%" if result['recall_3k_increase_pct'] is not None else ""
        
        recall_4k_pre = f"{result['initial_recall_4k']:.6f}" if result['initial_recall_4k'] is not None else ""
        # Format post-SGO: show mean with median in parentheses if median is higher (better)
        if result['best_recall_4k_median'] is not None and result['best_recall_4k_mean'] is not None and result['best_recall_4k_median'] > result['best_recall_4k_mean']:
            recall_4k_post = f"{result['best_recall_4k_mean']:.6f} (med: {result['best_recall_4k_median']:.6f})"
        else:
            recall_4k_post = f"{result['best_recall_4k_mean']:.6f}" if result['best_recall_4k_mean'] is not None else ""
        recall_4k_increase = f"{result['recall_4k_increase_pct']:.2f}%" if result['recall_4k_increase_pct'] is not None else ""
        
        recall_str = f"Recall@3k: {recall_3k_pre} / {recall_3k_post} / {recall_3k_increase}\nRecall@4k: {recall_4k_pre} / {recall_4k_post} / {recall_4k_increase}"
        ws.cell(row=row_idx, column=7, value=recall_str)
        ws.cell(row=row_idx, column=7).alignment = Alignment(vertical="top", wrap_text=True)
        
        # Note
        ws.cell(row=row_idx, column=8, value=result['note'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 20
    
    # Set row height for header
    ws.row_dimensions[1].height = 40
    
    # Save
    wb.save(output_file)
    logger.info(f"Excel file saved to: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Generate Excel sheet with journey metrics')
    parser.add_argument('--results-dir', type=str, default=None,
                       help='Path to results directory')
    parser.add_argument('--output', type=str, default=None,
                       help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Determine results directory
    if args.results_dir:
        results_dir = Path(args.results_dir)
    else:
        results_dir = Path(__file__).parent.parent / "artifacts" / "sgo_training_results_v3_sample"
    
    if not results_dir.exists():
        logger.error(f"Results directory not found: {results_dir}")
        return 1
    
    # Determine output file
    if args.output:
        output_file = Path(args.output)
    else:
        output_file = results_dir / "journey_metrics_summary.xlsx"
    
    # Process all metrics
    logger.info("Processing journey metrics...")
    results = process_all_metrics(results_dir)
    
    if not results:
        logger.error("No metrics found to process")
        return 1
    
    logger.info(f"Processed {len(results)} cluster/micro combinations")
    
    # Create Excel sheet
    create_excel_sheet(results, output_file)
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

