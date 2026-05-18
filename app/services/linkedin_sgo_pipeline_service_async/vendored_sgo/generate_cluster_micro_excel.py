#!/usr/bin/env python3
"""
Generate Excel sheet with all clusters and micro clusters metrics.
Includes WD, JSD, Behavior Loss, and Recall@K metrics.
"""

import json
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import sys
import argparse
import logging
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def load_delta_and_journey_metrics(results_dir: Path, cluster_id: str, micro_id: str) -> Dict[str, Optional[float]]:
    """
    Load theme_delta and text_delta from delta and journey files.
    
    Returns:
        Dictionary with initial_theme_delta_mean, best_theme_delta, 
        initial_text_delta_mean, best_text_delta
    """
    result = {
        'initial_theme_delta_mean': None,
        'best_theme_delta': None,
        'initial_text_delta_mean': None,
        'best_text_delta': None
    }
    
    # Load initial deltas from _delta directory
    delta_file = results_dir / "_delta" / cluster_id / f"{micro_id}_all_reviews_deltas.json"
    initial_theme_deltas = []
    initial_text_deltas = []
    
    if delta_file.exists():
        try:
            with open(delta_file, 'r', encoding='utf-8') as f:
                delta_data = json.load(f)
            
            for delta_entry in delta_data.get('deltas', []):
                theme_delta = delta_entry.get('theme_delta')
                text_delta = delta_entry.get('text_delta')
                
                if theme_delta is not None:
                    try:
                        initial_theme_deltas.append(float(theme_delta))
                    except (ValueError, TypeError):
                        pass
                
                if text_delta is not None:
                    try:
                        initial_text_deltas.append(float(text_delta))
                    except (ValueError, TypeError):
                        pass
        except Exception as e:
            logger.debug(f"Error loading delta file {delta_file}: {e}")
    
    # Calculate initial means
    if initial_theme_deltas:
        result['initial_theme_delta_mean'] = float(np.mean(initial_theme_deltas))
    if initial_text_deltas:
        result['initial_text_delta_mean'] = float(np.mean(initial_text_deltas))
    
    # Load journey file to find best deltas
    journey_file = results_dir / "_journey" / cluster_id / f"{micro_id}_journey.json"
    best_theme_deltas = []
    best_text_deltas = []
    
    if journey_file.exists():
        try:
            with open(journey_file, 'r', encoding='utf-8') as f:
                journey_data = json.load(f)
            
            for review_key, review_data in journey_data.items():
                # Get initial deltas
                initial_deltas = review_data.get('initial_prediction_deltas', {})
                initial_theme = initial_deltas.get('theme_delta')
                initial_text = initial_deltas.get('text_delta')
                
                # Start with initial as best
                best_theme = initial_theme
                best_text = initial_text
                
                # Check correction journey for better deltas
                correction_journey = review_data.get('correction_journey', [])
                for correction in correction_journey:
                    # Journey files use 'new_theme_delta' and 'new_text_delta'
                    correction_theme = correction.get('new_theme_delta')
                    correction_text = correction.get('new_text_delta')
                    
                    # Lower is better for deltas
                    if correction_theme is not None and (best_theme is None or correction_theme < best_theme):
                        best_theme = correction_theme
                    if correction_text is not None and (best_text is None or correction_text < best_text):
                        best_text = correction_text
                
                # Store best deltas for this review
                if best_theme is not None:
                    try:
                        best_theme_deltas.append(float(best_theme))
                    except (ValueError, TypeError):
                        pass
                if best_text is not None:
                    try:
                        best_text_deltas.append(float(best_text))
                    except (ValueError, TypeError):
                        pass
        except Exception as e:
            logger.debug(f"Error loading journey file {journey_file}: {e}")
    
    # Calculate best deltas (minimum means)
    if best_theme_deltas:
        result['best_theme_delta'] = float(np.mean(best_theme_deltas))
    if best_text_deltas:
        result['best_text_delta'] = float(np.mean(best_text_deltas))
    
    return result


def load_micro_cluster_metrics(file_path: Path, results_dir: Path) -> Optional[Dict[str, Any]]:
    """Load metrics from a micro cluster file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Extract cluster and micro IDs
        cluster_match = re.search(r'cluster_(\d+)', str(file_path))
        micro_match = re.search(r'micro_(\d+)', str(file_path))
        cluster_id = f"cluster_{cluster_match.group(1)}" if cluster_match else "unknown"
        micro_id = f"micro_{micro_match.group(1)}" if micro_match else "unknown"
        
        # Get persona name
        persona_name = data.get('persona_name', '')
        if not persona_name:
            metadata = data.get('metadata', {})
            persona_name = metadata.get('persona_name', '')
        
        tribe_metrics = data.get('tribe_metrics', {})
        if not tribe_metrics:
            return None
        
        # Load theme_delta and text_delta from delta/journey files
        delta_journey_metrics = load_delta_and_journey_metrics(results_dir, cluster_id, micro_id)
        
        # Calculate improvements
        initial_theme_delta = delta_journey_metrics['initial_theme_delta_mean']
        best_theme_delta = delta_journey_metrics['best_theme_delta']
        initial_text_delta = delta_journey_metrics['initial_text_delta_mean']
        best_text_delta = delta_journey_metrics['best_text_delta']
        
        # Calculate percentage improvements
        theme_improvement = None
        if initial_theme_delta is not None and best_theme_delta is not None and initial_theme_delta > 0:
            theme_improvement = ((initial_theme_delta - best_theme_delta) / initial_theme_delta) * 100.0
        
        text_improvement = None
        if initial_text_delta is not None and best_text_delta is not None and initial_text_delta > 0:
            text_improvement = ((initial_text_delta - best_text_delta) / initial_text_delta) * 100.0
        
        return {
            'cluster_id': cluster_id,
            'micro_id': micro_id,
            'persona_name': persona_name,
            'total_reviews': tribe_metrics.get('total_reviews', 0),
            'initial_wd_mean': tribe_metrics.get('initial_wd', {}).get('mean'),
            'post_wd_mean': tribe_metrics.get('wd', {}).get('mean'),
            'initial_jsd_mean': tribe_metrics.get('initial_jsd', {}).get('mean'),
            'post_jsd_mean': tribe_metrics.get('jsd', {}).get('mean'),
            'initial_recall_3k_mean': tribe_metrics.get('initial_recall_3k', {}).get('mean'),
            'post_recall_3k_mean': tribe_metrics.get('recall_3k', {}).get('mean'),
            'initial_recall_4k_mean': tribe_metrics.get('initial_recall_4k', {}).get('mean'),
            'post_recall_4k_mean': tribe_metrics.get('recall_4k', {}).get('mean'),
            'initial_deltas_mean': tribe_metrics.get('initial_deltas', {}).get('mean'),
            'best_delta': tribe_metrics.get('best_delta'),
            'improvement_in_wd_mean': tribe_metrics.get('improvement_in_wd', {}).get('mean'),
            'improvement_in_jsd_mean': tribe_metrics.get('improvement_in_jsd', {}).get('mean'),
            'improvement_in_recall_3k_mean': tribe_metrics.get('improvement_in_recall_3k', {}).get('mean'),
            'improvement_in_recall_4k_mean': tribe_metrics.get('improvement_in_recall_4k', {}).get('mean'),
            # New metrics for thematic and semantic loss
            'initial_theme_delta_mean': initial_theme_delta,
            'best_theme_delta': best_theme_delta,
            'theme_improvement_pct': theme_improvement,
            'initial_text_delta_mean': initial_text_delta,
            'best_text_delta': best_text_delta,
            'text_improvement_pct': text_improvement,
        }
    except Exception as e:
        logger.warning(f"Error loading {file_path}: {e}")
        return None


def get_best_overall_delta_from_journey(results_dir: Path, cluster_id: str, micro_id: str) -> Optional[float]:
    """
    Get the actual best (minimum) overall_delta from journey file corrections.
    This is the true best delta achieved during SGO, not just the minimum initial delta.
    """
    journey_file = results_dir / "_journey" / cluster_id / f"{micro_id}_journey.json"
    
    if not journey_file.exists():
        return None
    
    try:
        with open(journey_file, 'r', encoding='utf-8') as f:
            journey_data = json.load(f)
        
        all_overall_deltas = []
        
        for review_key, review_data in journey_data.items():
            # Get initial overall_delta
            initial_deltas = review_data.get('initial_prediction_deltas', {})
            initial_overall = initial_deltas.get('overall_delta')
            if initial_overall is not None:
                all_overall_deltas.append(float(initial_overall))
            
            # Get all correction overall_deltas
            correction_journey = review_data.get('correction_journey', [])
            for correction in correction_journey:
                new_overall = correction.get('new_overall_delta')
                if new_overall is not None:
                    all_overall_deltas.append(float(new_overall))
        
        if all_overall_deltas:
            return float(np.min(all_overall_deltas))
    except Exception as e:
        logger.debug(f"Error loading journey file {journey_file}: {e}")
    
    return None


def calculate_behavior_loss_percentage(initial_delta_mean: Optional[float], best_delta: Optional[float]) -> Optional[float]:
    """
    Calculate Behavior Loss (Thematic Loss % Diff).
    Formula: (Thematic Start - Thematic Loss End) / Theme Loss Start
    Where:
    - Thematic Start = initial_deltas_mean (average initial overall_delta)
    - Thematic Loss End = best_delta (minimum overall_delta achieved)
    """
    if initial_delta_mean is None or best_delta is None:
        return None
    
    if initial_delta_mean == 0:
        return None
    
    # Behavior Loss % = (initial - best) / initial * 100
    behavior_loss_pct = ((initial_delta_mean - best_delta) / initial_delta_mean) * 100.0
    
    return behavior_loss_pct


def format_value(value: Optional[float], decimals: int = 6) -> str:
    """Format a float value or return empty string if None."""
    if value is None or (isinstance(value, float) and (np.isnan(value) or np.isinf(value))):
        return ""
    return f"{value:.{decimals}f}"


def format_percentage(value: Optional[float], decimals: int = 2) -> str:
    """Format a percentage value or return empty string if None."""
    if value is None or (isinstance(value, float) and (np.isnan(value) or np.isinf(value))):
        return ""
    return f"{value:.{decimals}f}%"


def generate_excel(results_dir: Path, output_file: Path):
    """Generate Excel file with all cluster and micro cluster metrics."""
    logger.info(f"Loading micro cluster files from: {results_dir}")
    
    # Load all micro cluster files
    all_metrics = []
    for cluster_dir in sorted(results_dir.glob('cluster_*')):
        if not cluster_dir.is_dir():
            continue
        
        for summary_file in sorted(cluster_dir.glob('micro_*_summary_*.json')):
            metrics = load_micro_cluster_metrics(summary_file, results_dir)
            if metrics:
                all_metrics.append(metrics)
    
    logger.info(f"Loaded metrics from {len(all_metrics)} micro clusters")
    
    # Group by cluster
    clusters_data = {}
    for metrics in all_metrics:
        cluster_id = metrics['cluster_id']
        if cluster_id not in clusters_data:
            clusters_data[cluster_id] = []
        clusters_data[cluster_id].append(metrics)
    
    # Sort micro clusters within each cluster
    for cluster_id in clusters_data:
        clusters_data[cluster_id].sort(key=lambda x: int(x['micro_id'].split('_')[1]) if x['micro_id'] != 'unknown' else 999)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cluster Micro Metrics"
    
    # Define styles (all plain white background)
    header_font = Font(bold=True, size=11)
    cluster_font = Font(bold=True, size=11)
    micro_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define columns
    columns = [
        "Train (45 Mins)",
        "WD (Pre-SGO/Post-SGO/ Reduction %)",
        "JSD (Pre-SGO/Post-SGO/ Reduction %)",
        "Thematic Loss (JSD Improvement %)",
        "Semantic Loss (Text Delta Improvement %)",
        "Behavior Loss (Overall Delta % Diff)",
        "Recall@max(3,k) (Pre-Post SGO)/ % Increase in Recall",
        "Recall@max(4,k) (Pre-Post SGO)/ % Increase in Recall",
        "Note"
    ]
    
    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 15
    
    current_row = 2
    
    # Write data for each cluster
    for cluster_id in sorted(clusters_data.keys(), key=lambda x: int(x.split('_')[1]) if x != 'unknown' else 999):
        micro_clusters = clusters_data[cluster_id]
        
        # Cluster header row
        cluster_cell = ws.cell(row=current_row, column=1, value=f"{cluster_id} <Name>")
        cluster_cell.font = cluster_font
        cluster_cell.border = border
        
        # Merge cells for cluster row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
        
        current_row += 1
        
        # Write each micro cluster
        for metrics in micro_clusters:
            micro_id = metrics['micro_id']
            persona_name = metrics['persona_name'] or f"Tribe {micro_id.split('_')[1]}"
            
            # Micro cluster name (no tagging)
            micro_name = f"{micro_id} {persona_name}"
            micro_cell = ws.cell(row=current_row, column=1, value=micro_name)
            micro_cell.font = micro_font
            micro_cell.border = border
            
            # WD (Pre-SGO/Post-SGO/Reduction %)
            initial_wd = metrics.get('initial_wd_mean')
            post_wd = metrics.get('post_wd_mean')
            # Calculate improvement directly: (initial - post) / initial * 100 (lower is better)
            wd_reduction = None
            if initial_wd is not None and post_wd is not None and initial_wd > 0:
                wd_reduction = ((initial_wd - post_wd) / initial_wd) * 100.0
            
            wd_str = f"{format_value(initial_wd, 6)}/{format_value(post_wd, 6)}/{format_percentage(wd_reduction)}"
            ws.cell(row=current_row, column=2, value=wd_str).border = border
            
            # JSD (Pre-SGO/Post-SGO/Reduction %)
            initial_jsd = metrics.get('initial_jsd_mean')
            post_jsd = metrics.get('post_jsd_mean')
            # Calculate improvement directly: (initial - post) / initial * 100 (lower is better)
            jsd_reduction = None
            if initial_jsd is not None and post_jsd is not None and initial_jsd > 0:
                jsd_reduction = ((initial_jsd - post_jsd) / initial_jsd) * 100.0
            
            jsd_str = f"{format_value(initial_jsd, 6)}/{format_value(post_jsd, 6)}/{format_percentage(jsd_reduction)}"
            ws.cell(row=current_row, column=3, value=jsd_str).border = border
            
            # Thematic Loss (JSD Improvement %) - same as Column C
            thematic_loss_str = format_percentage(jsd_reduction)
            ws.cell(row=current_row, column=4, value=thematic_loss_str).border = border
            
            # Semantic Loss (Text Delta Improvement %)
            initial_text_delta = metrics.get('initial_text_delta_mean')
            best_text_delta = metrics.get('best_text_delta')
            # Calculate improvement directly: (initial - best) / initial * 100 (lower is better)
            text_improvement = None
            if initial_text_delta is not None and best_text_delta is not None and initial_text_delta > 0:
                text_improvement = ((initial_text_delta - best_text_delta) / initial_text_delta) * 100.0
            semantic_loss_str = format_percentage(text_improvement)
            ws.cell(row=current_row, column=5, value=semantic_loss_str).border = border
            
            # Behavior Loss (Overall Delta % Diff)
            # Get actual best delta from journey file (not just min initial delta)
            actual_best_delta = get_best_overall_delta_from_journey(
                results_dir, 
                metrics.get('cluster_id'), 
                metrics.get('micro_id')
            )
            # Fall back to stored best_delta if journey file not found
            best_delta_to_use = actual_best_delta if actual_best_delta is not None else metrics.get('best_delta')
            
            behavior_loss = calculate_behavior_loss_percentage(
                metrics.get('initial_deltas_mean'),
                best_delta_to_use
            )
            behavior_loss_str = format_percentage(behavior_loss) if behavior_loss is not None else ""
            ws.cell(row=current_row, column=6, value=behavior_loss_str).border = border
            
            # Recall@max(3,k) (Pre-Post SGO)/ % Increase in Recall
            initial_recall_3k = metrics.get('initial_recall_3k_mean')
            post_recall_3k = metrics.get('post_recall_3k_mean')
            # Calculate improvement directly: (post - initial) / initial * 100 (higher is better)
            recall_3k_increase = None
            if initial_recall_3k is not None and post_recall_3k is not None:
                if initial_recall_3k > 0:
                    recall_3k_increase = ((post_recall_3k - initial_recall_3k) / initial_recall_3k) * 100.0
                elif initial_recall_3k == 0 and post_recall_3k > 0:
                    # Special case: went from 0 to positive
                    recall_3k_increase = 1000.0  # Large fixed value to represent significant improvement
                else:
                    recall_3k_increase = 0.0
            
            recall_3k_str = f"{format_value(initial_recall_3k, 6)}/{format_value(post_recall_3k, 6)}/{format_percentage(recall_3k_increase)}"
            ws.cell(row=current_row, column=7, value=recall_3k_str).border = border
            
            # Recall@max(4,k) (Pre-Post SGO)/ % Increase in Recall
            initial_recall_4k = metrics.get('initial_recall_4k_mean')
            post_recall_4k = metrics.get('post_recall_4k_mean')
            # Calculate improvement directly: (post - initial) / initial * 100 (higher is better)
            recall_4k_increase = None
            if initial_recall_4k is not None and post_recall_4k is not None:
                if initial_recall_4k > 0:
                    recall_4k_increase = ((post_recall_4k - initial_recall_4k) / initial_recall_4k) * 100.0
                elif initial_recall_4k == 0 and post_recall_4k > 0:
                    # Special case: went from 0 to positive
                    recall_4k_increase = 1000.0  # Large fixed value to represent significant improvement
                else:
                    recall_4k_increase = 0.0
            
            recall_4k_str = f"{format_value(initial_recall_4k, 6)}/{format_value(post_recall_4k, 6)}/{format_percentage(recall_4k_increase)}"
            ws.cell(row=current_row, column=8, value=recall_4k_str).border = border
            
            # Note (empty for now, can add notes if needed)
            ws.cell(row=current_row, column=9, value="").border = border
            
            current_row += 1
        
        # Add empty row between clusters
        current_row += 1
    
    # Save workbook
    wb.save(output_file)
    logger.info(f"Excel file saved to: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Generate Excel sheet with cluster and micro cluster metrics',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('--results-dir', type=str, default=None,
                       help='Path to sgo_training_results directory (default: script directory)')
    parser.add_argument('--output', type=str, default=None,
                       help='Output Excel file path (default: cluster_micro_metrics.xlsx in results dir)')
    
    args = parser.parse_args()
    
    # Determine results directory
    if args.results_dir:
        results_dir = Path(args.results_dir)
    else:
        results_dir = Path(__file__).parent.parent / "artifacts" / "sgo_training_results_v6"
    
    if not results_dir.exists():
        logger.error(f"ERROR: Results directory not found: {results_dir}")
        return 1
    
    # Determine output file
    if args.output:
        output_file = Path(args.output)
    else:
        output_file = results_dir / "cluster_micro_metrics.xlsx"
    
    logger.info("=" * 80)
    logger.info("GENERATING EXCEL SHEET FOR CLUSTER AND MICRO CLUSTER METRICS")
    logger.info("=" * 80)
    logger.info(f"Results directory: {results_dir}")
    logger.info(f"Output file: {output_file}")
    
    try:
        generate_excel(results_dir, output_file)
        logger.info("=" * 80)
        logger.info("✓ Excel file generated successfully!")
        logger.info("=" * 80)
        return 0
    except Exception as e:
        logger.error(f"ERROR: Failed to generate Excel file: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    exit(main())

