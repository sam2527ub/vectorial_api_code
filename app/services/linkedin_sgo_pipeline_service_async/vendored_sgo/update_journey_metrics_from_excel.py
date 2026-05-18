#!/usr/bin/env python3
"""
Update journey metrics Excel by reading updated Pre-SGO/Post-SGO values
and recalculating improvements and losses.
"""

import re
from pathlib import Path
from typing import Optional, Tuple
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def parse_pre_post_value(cell_value: str) -> Tuple[Optional[float], Optional[float]]:
    """Parse Pre-SGO and Post-SGO values from cell string like '0.5571/0.4254/23.64%'."""
    if not cell_value:
        return None, None
    
    cell_str = str(cell_value).strip()
    
    # Handle different formats:
    # "0.5571/0.4254/23.64%" or "0.557106 / 0.535447 (med: 0.525455) / 5.68%"
    # Split by '/' and take first two parts
    parts = [p.strip() for p in cell_str.split('/')]
    
    if len(parts) >= 2:
        try:
            # First part is pre-SGO
            pre = float(parts[0].strip())
            
            # Second part is post-SGO (may have median info)
            post_str = parts[1].strip()
            # Remove median info if present: "0.4254 (med: 0.347091)" -> "0.4254"
            post_str = re.sub(r'\s*\(med:.*?\)', '', post_str)
            # Remove any trailing percentage or other text
            post_str = re.sub(r'%.*$', '', post_str).strip()
            post = float(post_str)
            return pre, post
        except (ValueError, AttributeError) as e:
            print(f"    Error parsing: {cell_str} - {e}")
            return None, None
    return None, None


def calculate_improvement_percentage(old_value: float, new_value: float) -> Optional[float]:
    """Calculate improvement as (old - new) / old * 100."""
    if old_value == 0:
        return None
    return ((old_value - new_value) / old_value) * 100.0


def update_excel_metrics(excel_file: Path):
    """Read Excel, parse updated values, recalculate metrics, and update."""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    print("Reading and recalculating metrics...")
    
    # Process each data row
    for row in range(2, ws.max_row + 1):
        train_col = ws.cell(row, 1).value
        if not train_col:
            continue
        
        print(f"\nProcessing: {train_col}")
        
        # Parse WD values
        wd_cell = ws.cell(row, 2).value
        wd_pre, wd_post = parse_pre_post_value(wd_cell)
        if wd_pre is not None and wd_post is not None:
            wd_reduction = calculate_improvement_percentage(wd_pre, wd_post)
            print(f"  WD: {wd_pre:.6f} -> {wd_post:.6f}, Reduction: {wd_reduction:.2f}%")
            # Update WD cell with recalculated reduction
            wd_str = f"{wd_pre:.4f}/{wd_post:.4f}/{wd_reduction:.2f}%" if wd_reduction is not None else f"{wd_pre:.4f}/{wd_post:.4f}/"
            ws.cell(row, 2, value=wd_str)
        
        # Parse JSD values
        jsd_cell = ws.cell(row, 3).value
        jsd_pre, jsd_post = parse_pre_post_value(jsd_cell)
        if jsd_pre is not None and jsd_post is not None:
            jsd_reduction = calculate_improvement_percentage(jsd_pre, jsd_post)
            print(f"  JSD: {jsd_pre:.6f} -> {jsd_post:.6f}, Reduction: {jsd_reduction:.2f}%")
            # Update JSD cell with recalculated reduction
            jsd_str = f"{jsd_pre:.6f} / {jsd_post:.6f} / {jsd_reduction:.2f}%" if jsd_reduction is not None else f"{jsd_pre:.6f} / {jsd_post:.6f} /"
            ws.cell(row, 3, value=jsd_str)
            
            # Update Thematic Loss (JSD Improvement %)
            thematic_loss = f"{jsd_reduction:.2f}%" if jsd_reduction is not None else ""
            ws.cell(row, 4, value=thematic_loss)
            print(f"  Thematic Loss: {thematic_loss}")
        
        # Get Semantic Loss and Behavior Loss (keep existing or recalculate if needed)
        semantic_loss_cell = ws.cell(row, 5).value
        behavior_loss_cell = ws.cell(row, 6).value
        
        # Parse existing percentages
        semantic_pct = None
        behavior_pct = None
        
        if semantic_loss_cell:
            semantic_str = str(semantic_loss_cell).replace('%', '').strip()
            try:
                semantic_pct = float(semantic_str)
            except ValueError:
                pass
        
        if behavior_loss_cell:
            behavior_str = str(behavior_loss_cell).replace('%', '').strip()
            try:
                behavior_pct = float(behavior_str)
            except ValueError:
                pass
        
        # Calculate Total Loss (sum of Thematic + Semantic + Behavior)
        total_loss = None
        if jsd_reduction is not None and semantic_pct is not None and behavior_pct is not None:
            total_loss = jsd_reduction + semantic_pct + behavior_pct
            print(f"  Total Loss: {total_loss:.2f}% (Thematic: {jsd_reduction:.2f}% + Semantic: {semantic_pct:.2f}% + Behavior: {behavior_pct:.2f}%)")
    
    # Add Total Loss column header if it doesn't exist
    if ws.max_column < 9:
        # Insert Total Loss column before Note column
        ws.insert_cols(8)
        header_cell = ws.cell(1, 8, value="Total Loss (%)")
        header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Move Note column to column 9
        for row in range(1, ws.max_row + 1):
            note_value = ws.cell(row, 9).value
            if note_value:
                ws.cell(row, 9).value = None
                ws.cell(row, 9, value=note_value)
    
    # Recalculate and fill Total Loss for all rows
    for row in range(2, ws.max_row + 1):
        # Get Thematic Loss
        thematic_cell = ws.cell(row, 4).value
        thematic_pct = None
        if thematic_cell:
            try:
                thematic_pct = float(str(thematic_cell).replace('%', '').strip())
            except ValueError:
                pass
        
        # Get Semantic Loss
        semantic_cell = ws.cell(row, 5).value
        semantic_pct = None
        if semantic_cell:
            try:
                semantic_pct = float(str(semantic_cell).replace('%', '').strip())
            except ValueError:
                pass
        
        # Get Behavior Loss
        behavior_cell = ws.cell(row, 6).value
        behavior_pct = None
        if behavior_cell:
            try:
                behavior_pct = float(str(behavior_cell).replace('%', '').strip())
            except ValueError:
                pass
        
        # Calculate Total Loss
        if thematic_pct is not None and semantic_pct is not None and behavior_pct is not None:
            total_loss = thematic_pct + semantic_pct + behavior_pct
            ws.cell(row, 8, value=f"{total_loss:.2f}%")
    
    # Adjust column widths
    ws.column_dimensions['H'].width = 15
    
    # Save
    wb.save(excel_file)
    print(f"\n✓ Excel file updated: {excel_file}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Update journey metrics Excel with recalculated values')
    parser.add_argument('--excel-file', type=str, required=True,
                       help='Path to Excel file to update')
    
    args = parser.parse_args()
    
    excel_file = Path(args.excel_file)
    if not excel_file.exists():
        print(f"Error: Excel file not found: {excel_file}")
        return 1
    
    update_excel_metrics(excel_file)
    return 0


if __name__ == "__main__":
    sys.exit(main())

